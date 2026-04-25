"""
Excel数据转换为分析图的程序
支持将Excel文件中的数据自动转换为多种分析图表，并在Excel中创建新工作表存放图表。
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import openpyxl
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart
from openpyxl.chart.series_factory import SeriesFactory
from openpyxl.chart.reference import Reference
from openpyxl.utils import get_column_letter
import os
import copy
import sys


class ExcelChartGenerator:
    """Excel图表生成器核心类"""
    
    def __init__(self):
        self.df = None
        self.file_path = None
        self.wb = None
        self.source_sheet = None
        self.chart_type_mapping = {
            "柱状图": "bar",
            "折线图": "line",
            "饼图": "pie",
            "散点图": "scatter"
        }
    
    def load_excel_file(self, file_path):
        """加载Excel文件"""
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")
        
        if not file_path.endswith(('.xlsx', '.xls')):
            raise ValueError("不支持的文件格式，请选择.xlsx或.xls文件")
        
        try:
            self.df = pd.read_excel(file_path)
            self.file_path = file_path
            
            if file_path.endswith('.xlsx'):
                self.wb = openpyxl.load_workbook(file_path)
            else:
                raise ValueError("仅支持.xlsx格式的图表生成（.xls只支持读取）")
            
            if self.df.empty:
                raise ValueError("Excel文件中没有数据")
            
            self.source_sheet = self.wb.active
            
            return True
        except (FileNotFoundError, ValueError) as e:
            raise e
        except Exception as e:
            raise Exception(f"读取Excel文件时出错: {str(e)}")
    
    def get_data_info(self):
        """获取数据信息"""
        if self.df is None:
            return None
        
        info = {
            "columns": list(self.df.columns),
            "shape": self.df.shape,
            "dtypes": {col: str(dtype) for col, dtype in self.df.dtypes.items()},
            "numeric_columns": list(self.df.select_dtypes(include=['number']).columns),
            "categorical_columns": list(self.df.select_dtypes(include=['object', 'category']).columns)
        }
        
        return info
    
    def create_chart_sheet(self, chart_configs):
        """创建包含图表的新工作表"""
        if self.df is None or self.wb is None:
            raise ValueError("没有加载数据")
        
        if "分析图表" in self.wb.sheetnames:
            self.wb.remove(self.wb["分析图表"])
        
        chart_sheet = self.wb.create_sheet("分析图表")
        
        source_name = self.source_sheet.title
        
        row_offset = 2
        for idx, config in enumerate(chart_configs):
            chart_type = config["type"]
            title = config["title"]
            category_col = config.get("category_col")
            data_col = config.get("data_col")
            x_col = config.get("x_col")
            y_col = config.get("y_col")
            
            chart_sheet.cell(row=row_offset, column=1, value=title)
            
            try:
                if chart_type == "bar":
                    self._create_bar_chart(chart_sheet, source_name, category_col, data_col, title, row_offset + 1)
                elif chart_type == "line":
                    self._create_line_chart(chart_sheet, source_name, category_col, data_col, title, row_offset + 1)
                elif chart_type == "pie":
                    self._create_pie_chart(chart_sheet, source_name, category_col, data_col, title, row_offset + 1)
                elif chart_type == "scatter":
                    self._create_scatter_chart(chart_sheet, source_name, x_col, y_col, title, row_offset + 1)
                
                row_offset += 25
            except Exception as e:
                chart_sheet.cell(row=row_offset + 1, column=1, value=f"创建图表时出错: {str(e)}")
                row_offset += 5
        
        return self.wb
    
    def _get_column_index(self, sheet, col_name):
        """获取列名对应的列索引"""
        for col_idx, cell in enumerate(sheet[1], 1):
            if cell.value == col_name:
                return col_idx
        raise ValueError(f"找不到列: {col_name}")
    
    def _create_bar_chart(self, chart_sheet, source_name, category_col, data_col, title, start_row):
        """创建柱状图"""
        chart = BarChart()
        chart.type = "col"
        chart.title = title
        chart.style = 10
        chart.y_axis.title = data_col
        chart.x_axis.title = category_col
        
        source_ws = self.wb[source_name]
        cat_idx = self._get_column_index(source_ws, category_col)
        data_idx = self._get_column_index(source_ws, data_col)
        
        max_row = len(self.df) + 1
        
        data_ref = Reference(source_ws, min_col=data_idx, min_row=1, max_col=data_idx, max_row=max_row)
        cat_ref = Reference(source_ws, min_col=cat_idx, min_row=2, max_row=max_row)
        
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cat_ref)
        chart.shape = 4
        
        chart_sheet.add_chart(chart, f"B{start_row}")
    
    def _create_line_chart(self, chart_sheet, source_name, category_col, data_col, title, start_row):
        """创建折线图"""
        chart = LineChart()
        chart.title = title
        chart.style = 10
        chart.y_axis.title = data_col
        chart.x_axis.title = category_col
        
        source_ws = self.wb[source_name]
        cat_idx = self._get_column_index(source_ws, category_col)
        data_idx = self._get_column_index(source_ws, data_col)
        
        max_row = len(self.df) + 1
        
        data_ref = Reference(source_ws, min_col=data_idx, min_row=1, max_col=data_idx, max_row=max_row)
        cat_ref = Reference(source_ws, min_col=cat_idx, min_row=2, max_row=max_row)
        
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cat_ref)
        chart.style = 13
        
        chart_sheet.add_chart(chart, f"B{start_row}")
    
    def _create_pie_chart(self, chart_sheet, source_name, category_col, data_col, title, start_row):
        """创建饼图"""
        chart = PieChart()
        chart.title = title
        chart.style = 10
        
        source_ws = self.wb[source_name]
        cat_idx = self._get_column_index(source_ws, category_col)
        data_idx = self._get_column_index(source_ws, data_col)
        
        max_row = len(self.df) + 1
        
        data_ref = Reference(source_ws, min_col=data_idx, min_row=1, max_col=data_idx, max_row=max_row)
        cat_ref = Reference(source_ws, min_col=cat_idx, min_row=2, max_row=max_row)
        
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cat_ref)
        
        chart_sheet.add_chart(chart, f"B{start_row}")
    
    def _create_scatter_chart(self, chart_sheet, source_name, x_col, y_col, title, start_row):
        """创建散点图"""
        chart = ScatterChart()
        chart.title = title
        chart.style = 10
        chart.y_axis.title = y_col
        chart.x_axis.title = x_col
        chart.scatterStyle = "marker"
        
        source_ws = self.wb[source_name]
        x_idx = self._get_column_index(source_ws, x_col)
        y_idx = self._get_column_index(source_ws, y_col)
        
        max_row = len(self.df) + 1
        
        x_values = Reference(source_ws, min_col=x_idx, min_row=2, max_row=max_row)
        y_values = Reference(source_ws, min_col=y_idx, min_row=2, max_row=max_row)
        y_title_ref = Reference(source_ws, min_col=y_idx, min_row=1, max_row=1)
        
        series = SeriesFactory(y_values, xvalues=x_values, title_from_data=True)
        series.marker.symbol = "circle"
        series.marker.size = 5
        series.graphicalProperties.line.noFill = True
        
        chart.series.append(series)
        
        chart_sheet.add_chart(chart, f"B{start_row}")
    
    def save_file(self, output_path=None):
        """保存文件"""
        if output_path is None:
            base, ext = os.path.splitext(self.file_path)
            output_path = f"{base}_分析图表{ext}"
        
        self.wb.save(output_path)
        return output_path


class ExcelChartApp:
    """Excel图表应用程序GUI"""
    
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Excel数据转换为分析图工具")
        self.root.geometry("900x650")
        self.root.resizable(True, True)
        
        self.generator = ExcelChartGenerator()
        self.chart_configs = []
        self.data_loaded = False
        
        self.setup_ui()
    
    def setup_ui(self):
        """设置用户界面"""
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        self.tab_file = ttk.Frame(self.notebook)
        self.tab_chart = ttk.Frame(self.notebook)
        self.tab_generate = ttk.Frame(self.notebook)
        
        self.notebook.add(self.tab_file, text="1. 文件选择")
        self.notebook.add(self.tab_chart, text="2. 图表配置")
        self.notebook.add(self.tab_generate, text="3. 生成图表")
        
        self.setup_file_tab()
        self.setup_chart_tab()
        self.setup_generate_tab()
    
    def setup_file_tab(self):
        """设置文件选择标签页"""
        frame = ttk.LabelFrame(self.tab_file, text="选择Excel文件", padding=10)
        frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        path_frame = ttk.Frame(frame)
        path_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Label(path_frame, text="文件路径:").pack(side=tk.LEFT, padx=5)
        self.file_path_var = tk.StringVar()
        file_entry = ttk.Entry(path_frame, textvariable=self.file_path_var, state="readonly")
        file_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill=tk.X, padx=5, pady=5)
        
        self.btn_browse = ttk.Button(btn_frame, text="浏览文件", command=self.browse_file)
        self.btn_browse.pack(side=tk.LEFT, padx=5)
        
        self.btn_load = ttk.Button(btn_frame, text="加载数据", command=self.load_data, state=tk.DISABLED)
        self.btn_load.pack(side=tk.LEFT, padx=5)
        
        info_container = ttk.Frame(frame)
        info_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        ttk.Label(info_container, text="数据信息:", font=("", 10, "bold")).pack(anchor=tk.W)
        
        self.info_text = tk.Text(info_container, height=20, state=tk.DISABLED, wrap=tk.WORD)
        self.info_text.pack(fill=tk.BOTH, expand=True, pady=5)
        
        scrollbar = ttk.Scrollbar(self.info_text, command=self.info_text.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.info_text.config(yscrollcommand=scrollbar.set)
    
    def setup_chart_tab(self):
        """设置图表配置标签页"""
        frame = ttk.LabelFrame(self.tab_chart, text="图表配置", padding=10)
        frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        config_frame = ttk.Frame(frame)
        config_frame.pack(fill=tk.X, padx=5, pady=5)
        
        params_frame = ttk.Frame(config_frame)
        params_frame.pack(fill=tk.X, padx=5, pady=5)
        
        row1 = ttk.Frame(params_frame)
        row1.pack(fill=tk.X, pady=2)
        
        ttk.Label(row1, text="图表类型:", width=10).pack(side=tk.LEFT, padx=5)
        self.chart_type_var = tk.StringVar()
        self.chart_type_combo = ttk.Combobox(row1, textvariable=self.chart_type_var, state="readonly", width=15)
        self.chart_type_combo['values'] = list(self.generator.chart_type_mapping.keys())
        self.chart_type_combo.pack(side=tk.LEFT, padx=5)
        
        ttk.Label(row1, text="图表标题:").pack(side=tk.LEFT, padx=(20, 5))
        self.chart_title_var = tk.StringVar(value="图表")
        self.chart_title_entry = ttk.Entry(row1, textvariable=self.chart_title_var, width=25)
        self.chart_title_entry.pack(side=tk.LEFT, padx=5)
        
        row2 = ttk.Frame(params_frame)
        row2.pack(fill=tk.X, pady=2)
        
        ttk.Label(row2, text="类别列/标签列:", width=15).pack(side=tk.LEFT, padx=5)
        self.category_col_var = tk.StringVar()
        self.category_col_combo = ttk.Combobox(row2, textvariable=self.category_col_var, state="readonly", width=20)
        self.category_col_combo.pack(side=tk.LEFT, padx=5)
        
        ttk.Label(row2, text="数值列:", width=10).pack(side=tk.LEFT, padx=(20, 5))
        self.data_col_var = tk.StringVar()
        self.data_col_combo = ttk.Combobox(row2, textvariable=self.data_col_var, state="readonly", width=20)
        self.data_col_combo.pack(side=tk.LEFT, padx=5)
        
        row3 = ttk.Frame(params_frame)
        row3.pack(fill=tk.X, pady=2)
        
        ttk.Label(row3, text="X轴列(散点图):", width=15).pack(side=tk.LEFT, padx=5)
        self.x_col_var = tk.StringVar()
        self.x_col_combo = ttk.Combobox(row3, textvariable=self.x_col_var, state="readonly", width=20)
        self.x_col_combo.pack(side=tk.LEFT, padx=5)
        
        ttk.Label(row3, text="Y轴列(散点图):", width=15).pack(side=tk.LEFT, padx=5)
        self.y_col_var = tk.StringVar()
        self.y_col_combo = ttk.Combobox(row3, textvariable=self.y_col_var, state="readonly", width=20)
        self.y_col_combo.pack(side=tk.LEFT, padx=5)
        
        hint_label = ttk.Label(params_frame, text="提示: 柱状图/折线图/饼图需要类别列和数值列，散点图需要X轴列和Y轴列", foreground="gray")
        hint_label.pack(fill=tk.X, pady=5)
        
        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill=tk.X, padx=5, pady=5)
        
        self.btn_add_chart = ttk.Button(btn_frame, text="添加图表配置", command=self.add_chart_config)
        self.btn_add_chart.pack(side=tk.LEFT, padx=5)
        
        self.btn_clear_configs = ttk.Button(btn_frame, text="清空所有配置", command=self.clear_configs)
        self.btn_clear_configs.pack(side=tk.LEFT, padx=5)
        
        list_container = ttk.Frame(frame)
        list_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        ttk.Label(list_container, text="已配置的图表列表:", font=("", 10, "bold")).pack(anchor=tk.W)
        
        self.config_listbox = tk.Listbox(list_container, height=12, font=("", 10))
        self.config_listbox.pack(fill=tk.BOTH, expand=True, pady=5)
        
        list_btn_frame = ttk.Frame(list_container)
        list_btn_frame.pack(fill=tk.X)
        
        self.btn_delete_config = ttk.Button(list_btn_frame, text="删除选中配置", command=self.delete_config)
        self.btn_delete_config.pack(side=tk.RIGHT, padx=5)
    
    def setup_generate_tab(self):
        """设置生成图表标签页"""
        frame = ttk.LabelFrame(self.tab_generate, text="生成图表", padding=10)
        frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        path_frame = ttk.Frame(frame)
        path_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Label(path_frame, text="输出路径:").pack(side=tk.LEFT, padx=5)
        self.output_path_var = tk.StringVar()
        output_entry = ttk.Entry(path_frame, textvariable=self.output_path_var, state="readonly")
        output_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill=tk.X, padx=5, pady=5)
        
        self.btn_select_output = ttk.Button(btn_frame, text="选择输出路径", command=self.select_output)
        self.btn_select_output.pack(side=tk.LEFT, padx=5)
        
        self.btn_generate = ttk.Button(btn_frame, text="生成图表", command=self.generate_charts, state=tk.DISABLED)
        self.btn_generate.pack(side=tk.LEFT, padx=5)
        
        hint_label = ttk.Label(frame, text="如果不选择输出路径，将在原文件同目录下生成带'_分析图表'后缀的文件", foreground="gray")
        hint_label.pack(fill=tk.X, padx=5, pady=5)
        
        result_container = ttk.Frame(frame)
        result_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        ttk.Label(result_container, text="生成日志:", font=("", 10, "bold")).pack(anchor=tk.W)
        
        self.result_text = tk.Text(result_container, height=15, state=tk.DISABLED, wrap=tk.WORD)
        self.result_text.pack(fill=tk.BOTH, expand=True, pady=5)
    
    def browse_file(self):
        """浏览选择Excel文件"""
        file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        
        if file_path:
            self.file_path_var.set(file_path)
            self.btn_load.config(state=tk.NORMAL)
    
    def load_data(self):
        """加载Excel数据"""
        file_path = self.file_path_var.get()
        
        if not file_path:
            messagebox.showerror("错误", "请先选择Excel文件")
            return
        
        try:
            self.generator.load_excel_file(file_path)
            data_info = self.generator.get_data_info()
            
            self.info_text.config(state=tk.NORMAL)
            self.info_text.delete(1.0, tk.END)
            
            self.info_text.insert(tk.END, "=" * 50 + "\n")
            self.info_text.insert(tk.END, "文件加载成功！\n")
            self.info_text.insert(tk.END, "=" * 50 + "\n\n")
            
            self.info_text.insert(tk.END, f"文件路径: {file_path}\n")
            self.info_text.insert(tk.END, f"数据形状: {data_info['shape'][0]} 行 × {data_info['shape'][1]} 列\n\n")
            
            self.info_text.insert(tk.END, "-" * 50 + "\n")
            self.info_text.insert(tk.END, "列信息:\n")
            self.info_text.insert(tk.END, "-" * 50 + "\n")
            for col in data_info['columns']:
                dtype = data_info['dtypes'][col]
                self.info_text.insert(tk.END, f"  {col} [{dtype}]\n")
            
            self.info_text.insert(tk.END, "\n" + "-" * 50 + "\n")
            self.info_text.insert(tk.END, f"数值列 ({len(data_info['numeric_columns'])}个):\n")
            self.info_text.insert(tk.END, "-" * 50 + "\n")
            for col in data_info['numeric_columns']:
                self.info_text.insert(tk.END, f"  ✓ {col}\n")
            
            self.info_text.insert(tk.END, "\n" + "-" * 50 + "\n")
            self.info_text.insert(tk.END, f"类别列 ({len(data_info['categorical_columns'])}个):\n")
            self.info_text.insert(tk.END, "-" * 50 + "\n")
            for col in data_info['categorical_columns']:
                self.info_text.insert(tk.END, f"  ✓ {col}\n")
            
            self.info_text.config(state=tk.DISABLED)
            
            columns = data_info['columns']
            self.category_col_combo['values'] = columns
            self.data_col_combo['values'] = columns
            self.x_col_combo['values'] = columns
            self.y_col_combo['values'] = columns
            
            if data_info['categorical_columns']:
                self.category_col_combo.set(data_info['categorical_columns'][0])
            elif columns:
                self.category_col_combo.set(columns[0])
                
            if data_info['numeric_columns']:
                self.data_col_combo.set(data_info['numeric_columns'][0])
                self.x_col_combo.set(data_info['numeric_columns'][0])
                if len(data_info['numeric_columns']) > 1:
                    self.y_col_combo.set(data_info['numeric_columns'][1])
                else:
                    self.y_col_combo.set(data_info['numeric_columns'][0])
            elif len(columns) > 1:
                self.data_col_combo.set(columns[1])
                self.x_col_combo.set(columns[0])
                self.y_col_combo.set(columns[1])
            
            self.data_loaded = True
            self.chart_configs = []
            self.config_listbox.delete(0, tk.END)
            
            self.btn_generate.config(state=tk.NORMAL)
            
            self.log_result("数据加载成功！可以开始配置图表。")
            
            messagebox.showinfo("成功", "数据加载成功！")
            
        except Exception as e:
            messagebox.showerror("错误", str(e))
            self.log_result(f"错误: {str(e)}")
    
    def add_chart_config(self):
        """添加图表配置"""
        if not self.data_loaded:
            messagebox.showerror("错误", "请先加载Excel数据")
            return
        
        chart_type = self.chart_type_var.get()
        title = self.chart_title_var.get()
        category_col = self.category_col_var.get()
        data_col = self.data_col_var.get()
        x_col = self.x_col_var.get()
        y_col = self.y_col_var.get()
        
        if not chart_type:
            messagebox.showerror("错误", "请选择图表类型")
            return
        
        if not title.strip():
            messagebox.showerror("错误", "请输入图表标题")
            return
        
        config = {
            "type": self.generator.chart_type_mapping[chart_type],
            "title": title.strip(),
            "category_col": category_col,
            "data_col": data_col,
            "x_col": x_col,
            "y_col": y_col
        }
        
        if chart_type in ["柱状图", "折线图", "饼图"]:
            if not category_col or not data_col:
                messagebox.showerror("错误", f"{chart_type}需要设置类别列和数值列")
                return
        elif chart_type == "散点图":
            if not x_col or not y_col:
                messagebox.showerror("错误", "散点图需要设置X轴列和Y轴列")
                return
        
        self.chart_configs.append(config)
        display_text = f"{len(self.chart_configs)}. {chart_type}: {title}"
        self.config_listbox.insert(tk.END, display_text)
        
        self.log_result(f"已添加图表配置: {display_text}")
    
    def clear_configs(self):
        """清空配置"""
        self.chart_configs = []
        self.config_listbox.delete(0, tk.END)
        self.log_result("已清空所有图表配置")
    
    def delete_config(self):
        """删除选中的配置"""
        selection = self.config_listbox.curselection()
        if selection:
            index = selection[0]
            config = self.chart_configs[index]
            self.config_listbox.delete(index)
            del self.chart_configs[index]
            self.log_result(f"已删除图表配置: {config['title']}")
        else:
            messagebox.showerror("错误", "请选择要删除的配置")
    
    def select_output(self):
        """选择输出路径"""
        file_path = filedialog.asksaveasfilename(
            title="选择输出文件路径",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        
        if file_path:
            self.output_path_var.set(file_path)
    
    def generate_charts(self):
        """生成图表"""
        if not self.chart_configs:
            messagebox.showerror("错误", "请先添加至少一个图表配置")
            return
        
        self.result_text.config(state=tk.NORMAL)
        self.result_text.delete(1.0, tk.END)
        
        self.log_result("=" * 50)
        self.log_result("开始生成图表...")
        self.log_result("=" * 50)
        self.log_result(f"图表配置数量: {len(self.chart_configs)}")
        
        for idx, config in enumerate(self.chart_configs, 1):
            self.log_result(f"  {idx}. {config['type']}: {config['title']}")
        
        self.log_result("")
        
        try:
            self.generator.create_chart_sheet(self.chart_configs)
            
            output_path = self.output_path_var.get()
            if not output_path:
                output_path = None
            
            result_path = self.generator.save_file(output_path)
            
            self.log_result("")
            self.log_result("=" * 50)
            self.log_result("✓ 图表生成成功！")
            self.log_result("=" * 50)
            self.log_result(f"输出文件: {result_path}")
            self.log_result(f"生成图表数: {len(self.chart_configs)}")
            self.log_result("")
            self.log_result("提示: 生成的图表已保存在Excel的'分析图表'工作表中")
            
            messagebox.showinfo("成功", f"图表生成成功！\n\n输出文件: {result_path}\n生成图表数: {len(self.chart_configs)}")
            
        except Exception as e:
            self.log_result("")
            self.log_result("=" * 50)
            self.log_result(f"✗ 错误: {str(e)}")
            self.log_result("=" * 50)
            messagebox.showerror("错误", str(e))
        
        self.result_text.config(state=tk.DISABLED)
    
    def log_result(self, message):
        """记录日志到结果文本框"""
        self.result_text.config(state=tk.NORMAL)
        self.result_text.insert(tk.END, message + "\n")
        self.result_text.see(tk.END)
        self.result_text.config(state=tk.DISABLED)
    
    def run(self):
        """运行应用程序"""
        self.root.mainloop()


def main():
    """主函数"""
    try:
        app = ExcelChartApp()
        app.run()
    except Exception as e:
        print(f"程序启动失败: {str(e)}")
        messagebox.showerror("错误", f"程序启动失败: {str(e)}")
        sys.exit(1)


if __name__ == "__main__":
    main()
