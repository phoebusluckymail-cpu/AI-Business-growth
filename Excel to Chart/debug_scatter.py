"""Check openpyxl chart internals"""
import openpyxl
from openpyxl.chart import ScatterChart
from openpyxl.chart.series import ScatterSeries
from openpyxl.chart.data_source import NumDataSource, NumData, NumVal
from openpyxl.chart.reference import Reference
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "Data"

ws.cell(row=1, column=1, value="X")
ws.cell(row=1, column=2, value="Y")
ws.cell(row=2, column=1, value=1)
ws.cell(row=2, column=2, value=2)
ws.cell(row=3, column=1, value=2)
ws.cell(row=3, column=2, value=4)
ws.cell(row=4, column=1, value=3)
ws.cell(row=4, column=2, value=6)

chart_sheet = wb.create_sheet("Charts")

chart = ScatterChart()
chart.title = "Test Scatter"
chart.y_axis.title = "Y"
chart.x_axis.title = "X"

x_values = Reference(ws, min_col=1, min_row=2, max_row=4)
y_values = Reference(ws, min_col=2, min_row=2, max_row=4)

print("Reference x_values:", x_values)
print("Reference y_values:", y_values)
print("Type:", type(x_values))

series = ScatterSeries()
series.marker.symbol = "circle"
series.marker.size = 5

print("\nBefore setting yVal/xVal:")
print("  series.yVal:", series.yVal)
print("  series.xVal:", series.xVal)

try:
    series.yVal = NumDataSource(numRef=y_values)
    print("  Set yVal with numRef: OK")
except Exception as e:
    print(f"  Set yVal with numRef: FAILED - {e}")

try:
    series.xVal = NumDataSource(numRef=x_values)
    print("  Set xVal with numRef: OK")
except Exception as e:
    print(f"  Set xVal with numRef: FAILED - {e}")

print("\nAfter setting yVal/xVal:")
print("  series.yVal:", series.yVal)
print("  series.xVal:", series.xVal)

chart.series.append(series)
chart_sheet.add_chart(chart, "B2")

print("\nCharts in sheet:", chart_sheet._charts)
print("Chart count:", len(chart_sheet._charts))

wb.save("test_scatter_debug.xlsx")
print("\nSaved to test_scatter_debug.xlsx")

import zipfile
import xml.etree.ElementTree as ET

with zipfile.ZipFile("test_scatter_debug.xlsx", 'r') as z:
    for name in z.namelist():
        if 'chart' in name.lower():
            print(f"\nFile: {name}")
            content = z.read(name).decode('utf-8')
            print(content[:2000])
