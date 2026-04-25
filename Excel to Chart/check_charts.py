"""检查生成的Excel文件中的图表"""
import openpyxl
import os

def check_excel_charts(file_path):
    print(f"检查文件: {file_path}")
    
    if not os.path.exists(file_path):
        print(f"文件不存在: {file_path}")
        return
    
    wb = openpyxl.load_workbook(file_path)
    
    print(f"工作表列表: {wb.sheetnames}")
    
    if "分析图表" in wb.sheetnames:
        ws = wb["分析图表"]
        print(f"\n'分析图表'工作表内容:")
        print(f"  维度: {ws.dimensions}")
        print(f"  最大行: {ws.max_row}")
        print(f"  最大列: {ws.max_column}")
        
        charts = ws._charts
        print(f"\n图表数量: {len(charts)}")
        
        for idx, chart in enumerate(charts, 1):
            print(f"\n图表 {idx}:")
            print(f"  类型: {type(chart).__name__}")
            print(f"  标题: {chart.title}")
            print(f"  位置: {chart.anchor}")
            print(f"  序列数: {len(chart.series)}")
            
            if hasattr(chart, 'x_axis'):
                print(f"  X轴标题: {chart.x_axis.title}")
            if hasattr(chart, 'y_axis'):
                print(f"  Y轴标题: {chart.y_axis.title}")
    else:
        print("没有找到'分析图表'工作表")

if __name__ == "__main__":
    test_files = [
        os.path.join(os.path.dirname(__file__), "测试输出_示例销售数据.xlsx"),
        os.path.join(os.path.dirname(__file__), "散点图测试输出.xlsx")
    ]
    
    for file_path in test_files:
        if os.path.exists(file_path):
            check_excel_charts(file_path)
            print("\n" + "=" * 60 + "\n")
