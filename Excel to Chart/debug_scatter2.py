"""Check openpyxl chart internals"""
import openpyxl
from openpyxl.chart import ScatterChart
from openpyxl.chart.series import ScatterSeries
from openpyxl.chart.data_source import NumDataSource, NumRef
from openpyxl.chart.reference import Reference
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Data"

ws.cell(row=1, column=1, value="X")
ws.cell(row=1, column=2, value="Y")
ws.cell(row=2, column=1, value=1)
ws.cell(row=2, column=2, value=2)
ws.cell(row=3, column=1, value=2)
ws.cell(row=3, column=2, value=4)
ws.cell(row=4, column=1, value=3)
ws.cell(row=4, column=2, value=6)

chart_sheet = wb.create_sheet("Charts")

chart = ScatterChart()
chart.title = "Test Scatter"
chart.y_axis.title = "Y"
chart.x_axis.title = "X"

max_row = 4

def make_num_ref(worksheet, col_idx, min_row, max_row):
    col_letter = get_column_letter(col_idx)
    sheet_name = worksheet.title.replace("'", "''")
    f = f"'{sheet_name}'!${col_letter}${min_row}:${col_letter}${max_row}"
    num_ref = NumRef()
    num_ref.f = f
    return num_ref

x_num_ref = make_num_ref(ws, 1, 2, max_row)
y_num_ref = make_num_ref(ws, 2, 2, max_row)

print("x_num_ref.f:", x_num_ref.f)
print("y_num_ref.f:", y_num_ref.f)

series = ScatterSeries()
series.marker.symbol = "circle"
series.marker.size = 5
series.graphicalProperties.line.noFill = True

series.yVal = NumDataSource(numRef=y_num_ref)
series.xVal = NumDataSource(numRef=x_num_ref)

chart.series.append(series)
chart_sheet.add_chart(chart, "B2")

wb.save("test_scatter_debug.xlsx")
print("\nSaved to test_scatter_debug.xlsx")

import zipfile
with zipfile.ZipFile("test_scatter_debug.xlsx", 'r') as z:
    for name in z.namelist():
        if 'chart' in name.lower():
            print(f"\nFile: {name}")
            content = z.read(name).decode('utf-8')
            print(content[:2000])
