"""创建示例Excel文件用于测试"""
import openpyxl
from openpyxl import Workbook
import os

wb = Workbook()
ws = wb.active
ws.title = "销售数据"

headers = ["月份", "产品A销售额", "产品B销售额", "产品C销售额", "客户数量", "满意度评分"]
for col, header in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=header)

data = [
    ["1月", 12000, 15000, 8000, 150, 4.5],
    ["2月", 13500, 16200, 9500, 165, 4.6],
    ["3月", 15000, 17800, 11000, 180, 4.7],
    ["4月", 16500, 19500, 12500, 200, 4.8],
    ["5月", 18000, 21000, 14000, 220, 4.9],
    ["6月", 19500, 22500, 15500, 240, 4.8],
    ["7月", 21000, 24000, 17000, 260, 4.9],
    ["8月", 22500, 25500, 18500, 280, 5.0],
    ["9月", 24000, 27000, 20000, 300, 4.9],
    ["10月", 25500, 28500, 21500, 320, 4.8],
    ["11月", 27000, 30000, 23000, 340, 4.9],
    ["12月", 28500, 31500, 24500, 360, 5.0],
]

for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

output_path = os.path.join(os.path.dirname(__file__), "示例销售数据.xlsx")
wb.save(output_path)
print(f"示例Excel文件已创建: {output_path}")
print(f"文件包含以下数据:")
print(f"  - 12个月的销售数据")
print(f"  - 3个产品的销售额")
print(f"  - 客户数量和满意度评分")
print(f"  - 可用于测试柱状图、折线图、饼图、散点图")
